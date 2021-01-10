import csv
import openpyxl


def load_csv(f, out, table_name, data_types, p):
    try:
        csv_file = open(f, 'r')
        csv_file.close()
    except FileNotFoundError:
        return 'File not found. ' \
               'Make sure to include the full file path if file is not in current active directory.'

    with open(f) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=",", quotechar=r'"')  # TODO - add click options

        cols = next(csv_reader)  # TODO - handling for inputs w/o a header row
        col_names = ", ".join(cols)
        vals = next(csv_reader)

        if data_types and len(data_types) != len(cols):
            data_types = None
            print('Warning: Number of datatypes != number of fields. '
                  'Data types inferred from first row of input file.')
        elif data_types:
            data_types = [convert_col_type(t) for t in data_types]
        else:
            data_types = [get_col_type(val) for val in vals]

        if p:
            print("CREATE TABLE {} (".format(table_name), end='')

            for i in range(len(data_types) - 1):
                print("{} {}, ".format(cols[i], data_types[i]), end='')

            print("{} {});".format(cols[-1], data_types[-1]))

            print("INSERT INTO {} ({}) VALUES ({});".format(table_name, col_names, ", ".join(vals)))

            for row in csv_reader:
                print("INSERT INTO {} ({}) VALUES ({});".format(table_name, col_names, ", ".join(row)))

        else:
            with open(out, "w") as f_out:
                f_out.write("CREATE TABLE {} (".format(table_name))
                for i in range(len(data_types) - 1):
                    f_out.write("{} {}, ".format(cols[i], data_types[i]))

                f_out.write("{} {}); \n".format(cols[-1], data_types[-1]))

                f_out.write("INSERT INTO {} ({}) VALUES ({});\n".format(table_name, col_names, ", ".join(vals)))

                for row in csv_reader:
                    f_out.write("INSERT INTO {} ({}) VALUES ({});\n".format(table_name, col_names, ", ".join(row)))


def load_excel(f, out, table_name, data_types, p):
    # Defaults to first sheet
    try:
        wb = openpyxl.load_workbook(f)
        ws = wb.active  # TODO - add click option
    except FileNotFoundError:
        return 'File not found. ' \
               'Make sure to include the full file path if file is not in current active directory.'

    cols = next(ws.values)  # TODO - add option to handle input files w/o a header
    col_names = ", ".join(cols)
    vals = [c.value for c in ws[1]]

    if data_types and len(data_types) != len(cols):
        data_types = None
        print('Warning: Number of datatypes != number of fields. '
              'Data types inferred from first row of input file.')
    elif data_types:
        data_types = [convert_col_type(t) for t in data_types]
    else:
        data_types = [get_col_type(val) for val in vals]

    if p:
        print("CREATE TABLE {} (".format(table_name), end='')

        for i in range(len(data_types) - 1):
            print("{} {}, ".format(cols[i], data_types[i]), end='')

        print("{} {});".format(cols[-1], data_types[-1]))

        print("INSERT INTO {} ({}) VALUES ({});".format(table_name, col_names, ", ".join(vals)))

        for row_cells in ws.iter_rows(min_row=3):
            print("INSERT INTO {} ({}) VALUES ({});".format(table_name, col_names, ", ".join([str(i.value) for i in row_cells])))
    else:
        with open(out, "w") as f_out:
            f_out.write("CREATE TABLE {} (".format(table_name))
            for i in range(len(data_types) - 1):
                f_out.write("{} {}, ".format(cols[i], data_types[i]))

            f_out.write("{} {}); \n".format(cols[-1], data_types[-1]))

            f_out.write("INSERT INTO {} ({}) VALUES ({});\n".format(table_name, col_names, ", ".join(vals)))

            for row_cells in ws.iter_rows(min_row=3):
                f_out.write("INSERT INTO {} ({}) VALUES ({});\n".format(table_name, col_names, ", ".join([str(i.value) for i in row_cells])))


def get_col_type(val):
    # only checks if value can be converted to int or float, leaves as str if type conversion errors
    # only checks the first instance in each column -- will break if later VALUES are in different formats

    if val.isnumeric() is True or val.strip("-").isnumeric() is True:
        return "NUMBER(22,0)"

    elif val.strip("-").strip(".").isnumeric() is True:
        return "NUMBER(22,2)"

    else:
        if len(val) >= 20:
            return "VARCHAR2(128)"
        else:
            return "VARCHAR2(35)"


def convert_col_type(type_str):
    if type_str == "str":
        return "VARCHAR2(35)"
    elif type_str == "int":
        return "NUMBER(22,0)"
    elif type_str == "float":
        return "NUMBER(22,2)"
    elif type_str[:3] == "str" and type_str[3:].isnumeric():
        return "VARCHAR(%s)" % type_str[3:]
    elif type_str[:3] == "int" and type_str[3:].isnumeric():
        return "NUMBER(%s,0)" % type_str[3:]
    else:
        print("Datatype %s is invalid." % type_str)
        raise NameError
